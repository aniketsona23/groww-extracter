import openpyxl
import fetcher
import scraper

obj = openpyxl.load_workbook("./Assignment.xlsx")
sheet1 = obj["Companies"]
sheet2 = obj["Assignement"]
companies = sheet1.cell(row=47,column=3).value
companies_list= list(map(lambda company: company.strip(" '"),companies[1:len(companies)-1].split(",")))
count= 0 


for i in range(10,20):
    company_search = scraper.gettting(companies_list[i-10])
    company_data = (fetcher.fetch_company_data(company_search))
    
    print("Company :",companies_list[i-10])
    print("P/E ratio :",company_data['peRatio'])
    print("P/b ratio :",company_data['pbRatio'])
    print("Debt ratio :",company_data['debtToEquity'])
    print("Promoters % :",company_data['promotersPerc'],'\n')
    
    sheet2.cell(row=i,column=5).value = companies_list[i-10].strip(" '")
    for j in range(7,11):
        sheet2.cell(row=i,column=j).value = [company_data['peRatio'],company_data['pbRatio'],company_data['debtToEquity'],company_data['promotersPerc']][j-7]

obj.save("./Assignment.xlsx")
